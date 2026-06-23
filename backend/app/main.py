from fastapi import FastAPI, UploadFile, File, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from app.api.routes import router
from app.database import engine, Base, SessionLocal
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Import des modules internes
from .transfo import (
    clean_and_correct_email,
    clean_and_correct_phone,
    load_formations_cache,
    correct_formation_and_campus,
    clean_and_correct_profil,
    has_error,
    get_first_value,
    VALID_DOMAINS
)
# from .postal_codes import get_city_from_postal_code
from .postal_codes import get_city_and_postal_code

from .rentree import calculate_rentree_date  # ← NOUVEAU

# Créer les tables dans la base de données
Base.metadata.create_all(bind=engine)

app = FastAPI(title="SA Plateforme API", version="1.0.0")

# Configuration CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.include_router(router, prefix="/api")

@app.get("/")
async def root():
    return {"message": "Bienvenue sur l'API SA Plateforme", "docs": "/docs"}

# Route pour l'aperçu avec corrections
@app.post("/process-and-preview")
async def process_and_preview(file: UploadFile = File(...)):
    """Traite le fichier et retourne les données corrigées au format JSON"""
    
    # Lire le fichier avec les bons types
    if file.filename.endswith('.csv'):
        df = pd.read_csv(file.file, dtype={'Téléphone': str, 'Phone number':str, 'zipcode': str})
    else:
        df = pd.read_excel(file.file, dtype={'Téléphone': str, 'Phone number':str, 'zipcode': str})
    df.columns = df.columns.str.strip()
    processed_rows = []
    stats = {
        'email_fixed': 0, 'phone_fixed': 0, 'city_fixed': 0, 
        'formation_fixed': 0, 'campus_fixed': 0, 'zip_fixed': 0, 
        'rentree_fixed': 0
    }
    db = SessionLocal()
    formations_cache = load_formations_cache(db)
    for index, row in df.iterrows():
        errors = []
        
        # Récupérer les valeurs (gérer les NaN)        
        raw_profil = get_first_value(row, ["Profiles", "profiles"])
        raw_nom = get_first_value(row, ["Nom", "nom", "Last name", "last name"])
        raw_prenom = get_first_value(row, ["Prénom", "prénom", "Prenom", "prenom", "First name", "first name"])
        raw_email = row.get("Email", "")
        raw_phone = get_first_value(row, ["Téléphone", "téléphone", "Telephone", "telephone", "Phone number", "phone number"])
        raw_zip = get_first_value(row, ["zipcode", "Zipcode", "Code postal", "code postal"])
        raw_city = get_first_value(row, ["city", "City", "Ville", "ville"])
        raw_formation = get_first_value(row, ["Souhaits de formations :", "Souhaits de formations", "souhaits de formations :", "souhaits de formations"])
        raw_campus = get_first_value(row, ["Choix de campus :","Choix de campus",])
        raw_classe = get_first_value(row, ["Actuellement, l'étudiant est en :","Actuellement, l’étudiant est en :","Actuellement, l'étudiant est en","Actuellement, l’étudiant est en",])
 
        # Convertir en string et gérer les NaN
        if pd.isna(raw_profil): raw_profil = ""
        if pd.isna(raw_nom): raw_nom = ""
        if pd.isna(raw_prenom): raw_prenom = ""
        if pd.isna(raw_email): raw_email = ""
        if pd.isna(raw_phone): raw_phone = ""
        if pd.isna(raw_zip): raw_zip = ""
        if pd.isna(raw_city): raw_city = ""
        if pd.isna(raw_formation): raw_formation = ""
        if pd.isna(raw_campus): raw_campus = ""
        if pd.isna(raw_classe): raw_classe = ""
        
        # Correction du profil
        corrected_profil, profil_valid, profil_msg = clean_and_correct_profil(raw_profil)
        if not profil_valid:
            errors.append({"field": "profil", "type": "error", "message": profil_msg})
        
        if not raw_nom:
            errors.append({"field": "nom", "type": "error", "message": "Nom manquant"})
            
        if not raw_prenom:
            errors.append({"field": "prénom", "type": "error", "message": "Préom manquant"})
        
        # Nettoyer les emails
        if isinstance(raw_email, str):
            raw_email = raw_email.strip()
        
        # Correction de l'email
        corrected_email, email_valid, email_msg = clean_and_correct_email(raw_email)
        if not email_valid:
            errors.append({"field": "email", "type": "error", "message": email_msg})
        elif email_msg:
            errors.append({"field": "email", "type": "warning", "message": email_msg})
            stats['email_fixed'] += 1

        # Correction du téléphone
        corrected_phone, phone_valid, phone_msg = clean_and_correct_phone(raw_phone)
        if not phone_valid:
            errors.append({"field": "telephone", "type": "error", "message": phone_msg})
        elif phone_msg:
            errors.append({"field": "telephone", "type": "warning", "message": phone_msg})
            stats['phone_fixed'] += 1
        
        # Correction code postal+ville
        error, corrected_zip, corrected_city, msg = get_city_and_postal_code(raw_zip,raw_city)
        if error==1:
            errors.append({"field": "codePostal", "type": "error", "message": msg})
            errors.append({"field": "ville", "type": "error", "message": msg})
        elif error==2:
            errors.append({"field": "ville", "type": "warning", "message": msg})
            stats['city_fixed'] += 1
        elif error==3:
            errors.append({"field": "codePostal", "type": "warning", "message": msg})
            errors.append({"field": "ville", "type": "error", "message": "Impossible de trouver une ville correspondante."})
        elif error==4:
            errors.append({"field": "codePostal", "type": "warning", "message": msg})
            stats['zip_fixed'] += 1
        elif error==5:
            errors.append({"field": "ville", "type": "error", "message": msg})
        elif error==6:
            errors.append({"field": "codePostal", "type": "error", "message": "Aucun code postale renseigné."})
            errors.append({"field": "ville", "type": "error", "message": "Aucune ville renseigné."})
        
        # Correction formation+campus    
        error, corrected_formation, corrected_campus, msg = correct_formation_and_campus(raw_formation,raw_campus,formations_cache)    
        if error==1:
            errors.append({"field": "campus", "type": "error", "message": msg})
                      
        if error==2:
            errors.append({"field": "formation", "type": "error", "message": msg})
                       
        if error==3:
            errors.append({"field": "formation", "type": "error", "message": msg})
        
        if error==4:
            errors.append({"field": "formation", "type": "error", "message": "Formation invalide."})
            errors.append({"field": "campus", "type": "error", "message": "Campus invalide."})
        
        # Calcul de la date de rentrée prévisionnelle (depuis rentree.py)
        rentree_date = None
        if raw_classe and str(raw_classe).strip():
            rentree_date = calculate_rentree_date(str(raw_classe))
            if rentree_date:
                stats['rentree_fixed'] += 1
                errors.append({
                    "field": "dateRentreePrev", 
                    "type": "warning", 
                    "message": f"Date de rentrée calculée: {rentree_date} (basé sur classe: {raw_classe})"
                })
        
        processed_rows.append({
            "id": index + 1,
            "profil": corrected_profil,
            "nom": str(raw_nom) if raw_nom else "",
            "prenom": str(raw_prenom) if raw_prenom else "",
            "email": corrected_email,
            "telephone": corrected_phone,
            "codePostal": corrected_zip,
            "ville": corrected_city,
            "formation": corrected_formation,
            "campus": corrected_campus,
            "classeActuelle": str(raw_classe) if raw_classe else "",
            "dateRentreePrev": rentree_date if rentree_date else "",
            "errors": errors
        })
    db.close()
    
    # Afficher les statistiques
    print("\n" + "="*50)
    print("STATISTIQUES DES CORRECTIONS")
    print("="*50)
    print(f"📧 Emails corrigés: {stats['email_fixed']}")
    print(f"📞 Téléphones corrigés: {stats['phone_fixed']}")
    print(f"📍 Codes postaux corrigés: {stats['zip_fixed']}")
    print(f"🏙️ Villes auto-remplies: {stats['city_fixed']}")
    print(f"🎓 Formations corrigées: {stats['formation_fixed']}")
    print(f"🏫 Campus corrigés: {stats['campus_fixed']}")
    print(f"📅 Dates rentrée calculées: {stats['rentree_fixed']}")
    print("="*50)
    
    salon_name = file.filename.replace('.xlsx', '').replace('.xls', '').replace('.csv', '')
    
    return JSONResponse(content={
        "success": True,
        "data": processed_rows,
        "total_rows": len(processed_rows),
        "salon_name": salon_name,
        "stats": stats
    })


# Route pour l'export Excel original
@app.post("/download")
# async def upload_file(data: dict):
async def upload_file(data: dict):
    rows = data.get("rows", [])
    error_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    warning_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    start_row = 2
    
    wb = load_workbook("Matrice import CRM.xlsx")
    ws = wb["Template"]
        
    for i, row in enumerate(rows, start=start_row):
        errors=row.get("errors",[])
        
        if has_error(errors, "profil"):
            ws.cell(row=i, column=3, value=row.get("profil","")).fill = error_fill
        else:
            ws.cell(row=i, column=3, value=row.get("profil",""))
            
        ws.cell(row=i, column=2, value=row.get("dateRentreePrev", ""))

        if has_error(errors, "nom"):
            ws.cell(row=i, column=5, value=row.get("nom", "")).fill = error_fill
        else:
            ws.cell(row=i, column=5, value=row.get("nom", ""))
        
        if has_error(errors, "prenom"):
            ws.cell(row=i, column=6, value=row.get("prenom", "")).fill = error_fill
        else:
            ws.cell(row=i, column=6, value=row.get("prenom", ""))
        
        if has_error(errors, "email"):
            ws.cell(row=i, column=7, value=row.get("email", "")).fill = error_fill
        else:
            ws.cell(row=i, column=7, value=row.get("email", ""))
        
        if has_error(errors, "telephone"):
            ws.cell(row=i, column=8, value=row.get("telephone", "")).fill = error_fill
        else:
            ws.cell(row=i, column=8, value=row.get("telephone", ""))
            
        if has_error(errors, "codePostal"):
            ws.cell(row=i, column=12, value=row.get("codePostal", "")).fill = error_fill
        else:
            ws.cell(row=i, column=12, value=row.get("codePostal", ""))
        
        if has_error(errors, "ville"):
            ws.cell(row=i, column=13, value=row.get("ville", "")).fill = error_fill
        else:
            ws.cell(row=i, column=13, value=row.get("ville", ""))
            
        ws.cell(row=i, column=15, value=row.get("classeActuelle", ""))
        
        ws.cell(row=i, column=22, value=row.get("campus", ""))
        
        if has_error(errors, "formation") and row.get("formation", "")!="":
            ws.cell(row=i, column=23, value=row.get("formation", "")).fill = error_fill
        else:
            ws.cell(row=i, column=23, value=row.get("formation", ""))
        


    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="test"'}
    )